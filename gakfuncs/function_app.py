import azure.functions as func
import logging
from ipaddress import IPv4Interface
import json
import re
import io
import pytz
import time
import datetime
from openpyxl import load_workbook, Workbook

app = func.FunctionApp(http_auth_level=func.AuthLevel.ANONYMOUS)

def hexip(ipadr):
  hx = []
  for i in str(ipadr).split('.'):
    hx.append(format(int(i), '02X'))
  return ':'.join(hx)


def convert_ip(ipadr):
  data = {}
  try:
    ip = IPv4Interface(ipadr)
    st, ed = (str(ip.network[0]), str(ip.network[-1]))
    val = str(ip.ip)
    netmask = str(str(ip.netmask))
    data['input'] = ipadr
    data['inputIP'] = val
    data['inputLong'] = ip._ip_int_from_string(val)
    data['inputHex'] = hexip(val)
    data['cidr'] = ip.with_prefixlen
    data['iprange'] = '%s-%s' % (st,ed)
    data['iprangeint'] = '%s-%s' % (ip._ip_int_from_string(st),ip._ip_int_from_string(ed))
    data['iprangehex'] = '%s-%s' % (hexip(st), hexip(ed))
    data['ipaddresses'] = ip.network.num_addresses
    data['maskbits'] = ip.with_prefixlen.split('/')[-1]
    data['netmask'] = netmask
    data['netmaskhex'] = hexip(netmask)
  except Exception as es:
    logging.error('Exception: %s' % es)
  return data


@app.route(route="subnet")
def subnet(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Python HTTP trigger subnet function processed a request.')

    ipaddr = None
    data = {}
    if 'q' in req.params:
        ipaddr = req.params.get('q', None)
    if ipaddr:
        data = convert_ip(ipaddr)
    # data['ip'] = req.remote_addr
    # data['host'] = request.host
    # return jsonify(data)
    return func.HttpResponse(
        json.dumps(data),
         mimetype="application/json"
        )


@app.route(route="epoch")
def epoch(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Python Http trigger epoch function processed a request.')
    try:
        e = req.params.get('e', None)
        if e is None:
            now = int(time.time())

        intz = req.params.get('intz', 'Asia/Kolkata')
        UTC = pytz.utc
        tz = pytz.timezone(intz)

        udt = datetime.datetime.fromtimestamp(now, UTC)
        dt = datetime.datetime.fromtimestamp(now, tz)

        format = '%A, %d %B %Y %H:%M:%S'
        udt_st = udt.strftime(format)
        dt_st = dt.strftime(format)

        fmt = '%Y:%m:%d %H:%M:%S %Z %z'
        offset = dt.strftime('%z')
        vals = re.match(r'([+,-])(\d\d)(\d\d)', offset)
        zone = ''
        if vals:
            zone = 'GMT %s%s:%s' % vals.groups()

        country = req.params.get('country', None)
        cdata = []
        if country:
            for tx in pytz.country_timezones[country]:
                st = pytz.timezone(tx)
                dtc = datetime.datetime.fromtimestamp(now, st)
                cdata.append(dtc.strftime(fmt))

        ret_val = {
        'epoch': now,
        'utc': udt_st,
        'current': '%s %s' % (dt_st, zone),
        'country': cdata
        }
        
        return func.HttpResponse(
        json.dumps(ret_val),
            mimetype='application/json'
        )
    except Exception as es:
        logging.error('Exception: %s' % es)
       

def get_reportdata(p, r, n):
    data = []
    MR = r/(12.0 *  100)
    val1 = pow((1+MR),n)
    val2 = pow((1+MR),n) - 1.0
    val = (p * MR * val1)/val2
    emi = int(val)
    tint = 0
    for i in range(int(n)):
        intr = p * MR
        tint = int(tint + intr)
        val = p - emi + intr
        row = {
          'month': i+1,
          'principal': int(p),
          'interest': int(intr),
          'total': tint,
          'balance': int(val),
          'emi': emi
        }
        p = int(val)
        data.append(row)
    return emi, data


def get_report_metadata():
    metadata = {
        'month': {'width': 10, 'format': 'text', 'title': 'Month'},
        'principal': {'width': 80, 'format': 'text', 'title': 'Principal'},
        'interest': {'width': 30, 'format': 'text', 'title': 'Interest'},
        'total': {'width': 70, 'format': 'text', 'title': 'Total Interest'},
        'balance': {'width': 70, 'format': 'text', 'title': 'Balance'},
        'emi': {'width': 30, 'format': 'text', 'title': 'EMI'},
        'order': ['month', 'principal', 'interest', 'total', 'balance', 'emi']
    }
    return metadata


def get_reportbytes(appdata):
    metadata = get_report_metadata()
    kwargs = {
        'file': '',
        'path': '',
        'meta': metadata,
        'appdata': appdata,
        'gz': '',
        'index': 0
    }
    wb = Workbook()
    ws = wb['Sheet']
    # Add column formatting here.
    for colid, col in enumerate(metadata['order']):
        ws.cell(row=1, column=colid + 1).value = metadata[col]['title']
    for rowid, record in enumerate(appdata):
        for colid, col in enumerate(metadata['order']):
            ws.cell(row=(rowid + 2), column=(colid + 1)).value = record[col]

    ioBytes = io.BytesIO()
    wb.save(ioBytes)
    ioBytes.seek(0)

    return ioBytes.getvalue()

@app.route(route="emi")
def emi(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Python Http emi trigger function processed a request.')
    p = req.params.get('p')
    r = req.params.get('r')
    n = req.params.get('n')
    t = req.params.get('t')
    try:
        emi, data = get_reportdata(int(p), int(r), int(n))
        if t and t == 'csv':
          csvdata = ['month,principal,interest,total,balance,emi',]
          for i in data:
            csvrow = "{:-3d},{:-6d},{:-5d},{:-5d},{:-6d},{:-5d}".format(i['month'],
                                                                        i['principal'], 
                                                                        i['interest'], 
                                                                        i['total'], 
                                                                        i['balance'], 
                                                                        i['emi'])
            csvdata.append(csvrow)
          return func.HttpResponse(
             '\n'.join(csvdata),
             mimetype='text/plain'
            )
        elif t and t == 'xlsx':
          content = get_reportbytes(data)
          return func.HttpResponse(
            content,
            headers={"Content-Disposition": 'attachment; filename="emi.xlsx"'},
            mimetype='application/vnd.ms-excel',
            status_code=200,
          )
        else:
          return func.HttpResponse(
            json.dumps(data),
            mimetype='application/json'
          )
    except Exception as es:
       logging.error('Exception: %s' %es)




